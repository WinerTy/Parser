from openpyxl import load_workbook


class ExcelParser:
    def __init__(self, file_path: str):
        self.file = load_workbook(file_path)
        self.sheet = self.file.active
        self.path = file_path

    @property
    def start_row(self) -> int:
        return self.find_start_row_with_products()

    @property
    def end_row(self) -> int:
        return self.find_end_row_with_products() - 1

    def find_start_row_with_products(self) -> int:
        """method for find start product table index

        Returns:
            int: index for start product table
        """
        start_row = 1

        if self.sheet.max_row == 0:
            return start_row

        for row_cells in self.sheet.iter_rows(min_row=1, max_col=2):
            if len(row_cells) > 1 and row_cells[1].value == "№":
                start_row = row_cells[1].row
                break

        return start_row

    def find_end_row_with_products(self) -> int:
        """method for find end product table index

        Returns:
            int: index for end product table
        """

        for r_idx in range(self.sheet.max_row, 0, -1):
            for c_idx in range(self.sheet.max_column, 0, -1):
                cell = self.sheet.cell(row=r_idx, column=c_idx)
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    if cell_text == "Итого:":
                        return r_idx - 1

        return self.sheet.max_row
