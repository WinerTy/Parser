from openpyxl import load_workbook
import pandas as pd

import numpy as np
from typing import List, Dict, Tuple
import re

file_path = "data/data.xlsx"


category_rule = {
    "Двигатель": {
        "Амортизаторы": ["амортизатор", "подвеска"],
        "Распредвал": ["втулка распределительного вала"],
        "Опоры": ["подушка двигателя"],
        "Коленчатый Вал": ["вал"],
        "Система Зажигания": ["зажигания", "свеча"],
    },
    "Тормозная Система": {
        "Барабаны": ["барабан"],
        "Втулки": ["втулка"],
        "Шланги": ["шланг"],
    },
    "Ходовая часть": {
        "Стабилизаторы": ["втулка-подушка", "подушка"],
        "Рессоры": ["рессоры"],
    },
    "Крепеж": {"Гайки": ["гайка"]},
}


class ExselParser:
    def __init__(self, file_path: str):
        self.file = load_workbook(file_path)
        self.sheet = self.file.active

    @property
    def start_row(self) -> int:
        return self.find_start_row_with_products()

    @property
    def end_row(self) -> int:
        return self.find_end_row_with_products()

    def find_start_row_with_products(self) -> int:
        """method for find start product table index

        Returns:
            int: index for start product table
        """
        start_row = 1

        if self.sheet.max_row == 0:
            return start_row

        for row_cells in self.sheet.iter_rows(min_row=1, max_col=2):
            if len(row_cells) > 1 and row_cells[1].value == "№":
                start_row = row_cells[1].row
                break

        return start_row

    def find_end_row_with_products(self) -> int:
        """method for find end product table index

        Returns:
            int: index for end product table
        """

        for r_idx in range(self.sheet.max_row, 0, -1):
            for c_idx in range(self.sheet.max_column, 0, -1):
                cell = self.sheet.cell(row=r_idx, column=c_idx)
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    if cell_text == "Итого:":
                        return r_idx

        return self.sheet.max_row


class ProductParser(ExselParser):
    def __init__(self, file_path: str):
        super().__init__(file_path)
        self.df = pd.read_excel(file_path, skiprows=self.start_row - 1)
        # Вынести отдельно
        self._delete_empty_columns()
        self.add_category_column()
        #####
        self.additional_table_columns = (
            self.delete_dublicate_title()
        )  # Массив индексов колонок внутри таблицы

    def delete_dublicate_title(self) -> List[int]:
        table_labels_title = self.df[self.df["№"] == "№"].index
        self.df.drop(axis=0, index=table_labels_title, inplace=True)
        self.df.dropna(axis=0, how="all", inplace=True)
        return table_labels_title

    def add_category_column(self) -> None:
        target = "Артикул"
        idx = self.df.columns.get_loc(target)
        self.df.insert(loc=idx + 1, column="Категория", value=np.nan)
        self.df.insert(loc=idx + 2, column="Подкатегория", value=np.nan)

    def _delete_empty_columns(self) -> None:
        cols_for_drop = []

        for col in self.df.columns:
            col_name = str(col)
            if col_name.startswith("Unnamed"):
                cols_for_drop.append(col_name)

        if cols_for_drop:
            self.df.drop(columns=cols_for_drop, inplace=True)

    def add_category_to_product(self, product_name: str):
        product_name = str(product_name).lower()

        for category, subcategories in category_rule.items():
            for subcategory, keywords in subcategories.items():
                for keyword in keywords:
                    # Простое вхождение подстроки:
                    # TODO улучшить алгоритм
                    if re.search(
                        r"\b" + re.escape(keyword.lower()) + r"\b", product_name
                    ):
                        return category, subcategory
        return "Не определена", "Не определена"


parser = ProductParser(file_path)


result = parser.df
result[["Категория", "Подкатегория"]] = result["Товары (работы, услуги)"].apply(
    lambda name: pd.Series(parser.add_category_to_product(name))
)
result.to_excel("data/result.xlsx", index=False)

result.to_csv("data/result.csv", index=False, sep=";")


print(parser.df.head(30))
category_not_assign = len(result[result["Категория"] == "Не определена"])
print("Категорий не определено: ", category_not_assign)


class CategoryMatcher:
    def __init__(self, category_rules: Dict):
        self.category_rules = category_rules
        self.keyword_map = self._build_keyword_map()
        self.unmatched_products = set()

    def _build_keyword_map(self) -> Dict[str, Tuple[str, str]]:
        """Создает обратный индекс ключевых слов"""
        keyword_map = {}
        for category, subcategories in self.category_rules.items():
            for subcategory, keywords in subcategories.items():
                for keyword in keywords:
                    keyword_lower = keyword.lower()
                    if keyword_lower not in keyword_map:
                        keyword_map[keyword_lower] = (category, subcategory)
        return keyword_map

    def _preprocess_text(self, text: str) -> str:
        """Предварительная обработка текста"""
        text = str(text).lower()
        text = re.sub(r"[^\w\s]", " ", text)  # Удаляем пунктуацию
        text = re.sub(r"\s+", " ", text).strip()  # Удаляем лишние пробелы
        return text

    def match_product(self, product_name: str) -> Tuple[str, str]:
        """Находит категорию и подкатегорию для продукта"""
        product_name = self._preprocess_text(product_name)
        words = product_name.split()

        # Проверяем сначала полные совпадения с ключевыми фразами
        for keyword, (category, subcategory) in self.keyword_map.items():
            if (
                len(keyword.split()) > 1
            ):  # Если ключевая фраза состоит из нескольких слов
                if keyword in product_name:
                    return category, subcategory

        # Затем проверяем отдельные слова
        for word in words:
            if word in self.keyword_map:
                return self.keyword_map[word]

        # Если не нашли, добавляем в список нераспознанных
        self.unmatched_products.add(product_name)
        return "Не определена", "Не определена"


matcher = CategoryMatcher(category_rule)


print(
    matcher.match_product(
        "Шланг тормозов гибкий (гайка-штуцер) (0,557м) ПАЗ Вектор NEXT"
    )
)
